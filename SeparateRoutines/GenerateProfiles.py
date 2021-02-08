import copy
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import re
from pathlib import Path
"""This is a routine to extract profiles from AHN data. It is based on detecting points with the highest covariance in slope, whcih are the crest, toe and berm lines. 
For some profiles manual point adjustments have been made. Therefore: handle this with care and check visually whether generated profiles are correct!!!"""

def extract_profile_data(profile_data, window, titel, path):
    # Interpolate gaps with no data
    profile_data['x'] = profile_data['x'].round(0).astype(int)
    profile_data = profile_data.sort_values('x')
    df = pd.DataFrame({'x': [i for i in range(profile_data['x'].iloc[0], profile_data['x'].iloc[-1], 2)]})
    df = pd.merge(df, profile_data, how='left', on='x').interpolate()

    # Smoothen of the profile and calculate its derivative and covariation
    df['z_averaged'] = df['z'].rolling(window=window, min_periods=1, center=True).mean()
    df['z_derivative'] = df['z_averaged'].diff()
    df['z_covariance'] = df['z'].rolling(window=window*2, min_periods=1, center=True).cov()

    # Identify the slopes from the covariance plot
    slopes = df[(df['z_covariance'] >= 0.15) & (df['x'] >= -50) & (df['x'] <= 50)].reset_index().rename(columns={'index': 'profile index'})

    # Correct the selected slopes on gradient direction to get rid of non-levee slopes
    slopes = slopes[((slopes['x'] < 0) & (slopes['z_derivative'] > 0)) | ((slopes['x'] > 0) & (slopes['z_derivative'] < 0))].reset_index(drop=True)
    pivot_index = slopes[slopes['profile index'].diff() > 1].reset_index().rename(columns={'index': 'slope index'})

    # Correct selected slopes on covariance intensity to get rid of small disturbances like ditches
    slopes_corrected = copy.deepcopy(slopes)
    for i in range(len(pivot_index)):
        if i == 0:
            first_slope_intensity = slopes.iloc[0:int(pivot_index.iloc[i]['slope index'])]['z_covariance'].max()
            if first_slope_intensity < 0.3:
                slopes_corrected = slopes_corrected.drop(slopes.iloc[0:int(pivot_index.iloc[i]['slope index'])].index)

        elif i != 0 and i != len(pivot_index)-1:
            intermediate_slope_intensity = slopes.iloc[int(pivot_index.iloc[i-1]['slope index']):int(pivot_index.iloc[i]['slope index'])]['z_covariance'].max()
            if intermediate_slope_intensity < 0.3:
                slopes_corrected = slopes_corrected.drop(slopes.iloc[int(pivot_index.iloc[i-1]['slope index']):int(pivot_index.iloc[i]['slope index'])].index)

        elif i != 0 and i == len(pivot_index)-1:
            intermediate_slope_intensity = slopes.iloc[int(pivot_index.iloc[i-1]['slope index']):int(pivot_index.iloc[i]['slope index'])]['z_covariance'].max()
            if intermediate_slope_intensity < 0.3:
                slopes_corrected = slopes_corrected.drop(slopes.iloc[int(pivot_index.iloc[i-1]['slope index']):int(pivot_index.iloc[i]['slope index'])].index)

            last_slope_intensity = slopes.iloc[int(pivot_index.iloc[len(pivot_index)-1]['slope index']-1):]['z_covariance'].max()
            if last_slope_intensity < 0.3:
                slopes_corrected = slopes_corrected.drop(slopes.iloc[int(pivot_index.iloc[len(pivot_index)-1]['slope index']-1):].index)

    # Overwrite with the corrected versions and calculate the pivot indexes of the profile
    slopes = slopes_corrected.reset_index(drop=True)
    pivot_index = slopes[slopes['profile index'].diff() > 1].reset_index().rename(columns={'index': 'slope index'})

    # Find the profile coordinates of the levee with respect to the outer_crest_index
    outer_crest_index = int(df[df['x'] == 0].index.values)
    matched_pivot_index = int((pivot_index['profile index'] - outer_crest_index).abs().idxmin())
    matched_slope_index = int(pivot_index.iloc[matched_pivot_index]['slope index'])

    # In case of no berm
    if matched_pivot_index == 0:
        inner_crest_index = int(slopes.iloc[pivot_index['slope index'].values[0]-1]['profile index'])
        inner_toe_index = int(slopes.iloc[0]['profile index'])
        if len(pivot_index) == 1:
            outer_toe_index = int(slopes.iloc[-1]['profile index'])
        else:
            outer_toe_index = int(slopes.iloc[int(pivot_index.iloc[matched_pivot_index+1]['slope index']-1)]['profile index'])

    # In case of a berm
    else:
        inner_crest_index = int(slopes.iloc[matched_slope_index-1]['profile index'])
        berm_inner_toe_side_index = int(slopes.iloc[int(pivot_index.iloc[matched_pivot_index-1]['slope index']-1)]['profile index'])
        berm_inner_crest_side_index = int(pivot_index.iloc[matched_pivot_index-1]['profile index'])
        if matched_pivot_index == pivot_index.index[-1]:
            outer_toe_index = int(slopes.iloc[-1]['profile index'])
        else:
            outer_toe_index = int(slopes.iloc[int(pivot_index.iloc[matched_pivot_index+1]['slope index']-1)]['profile index'])
        if matched_pivot_index-1 == pivot_index.index[0]:
            inner_toe_index = int(slopes.iloc[0]['profile index'])
        else:
            inner_toe_index = int(pivot_index.iloc[matched_pivot_index-2]['profile index'])

    # Manual corrections
    cross_section = int(re.findall('\d+', titel)[0])
    if cross_section == 1: inner_toe_index = 85
    if cross_section == 6: inner_toe_index = 83
    if cross_section == 15: inner_toe_index = 83
    if cross_section == 27: inner_toe_index = 84
    if cross_section == 33: matched_pivot_index = 0
    if cross_section == 39: inner_toe_index = 83
    if cross_section == 47: inner_toe_index = 82
    if cross_section == 53: inner_toe_index = 85
    if cross_section == 58: inner_toe_index = 81
    if cross_section == 59: inner_toe_index = 82
    if cross_section == 60: inner_toe_index = 84
    if cross_section == 64: inner_toe_index = 78
    if cross_section == 65: inner_toe_index = 67
    if cross_section == 74: inner_toe_index = 83
    if cross_section == 83: matched_pivot_index = 0

    # Extract x and z coordinates of the profile coordinates
    if matched_pivot_index == 0:
        x_coords = [df.loc[inner_toe_index]['x'], df.loc[inner_crest_index]['x'], df.loc[outer_crest_index]['x'], df.loc[outer_toe_index]['x']]
        z_coords = [df.loc[inner_toe_index]['z'], (df.loc[inner_crest_index]['z'] + df.loc[outer_crest_index]['z']) / 2, (df.loc[inner_crest_index]['z'] + df.loc[outer_crest_index]['z']) / 2, df.loc[outer_toe_index]['z']]
    else:
        x_coords = [df.loc[inner_toe_index]['x'], df.loc[berm_inner_toe_side_index]['x'], df.loc[berm_inner_crest_side_index]['x'], df.loc[inner_crest_index]['x'], df.loc[outer_crest_index]['x'], df.loc[outer_toe_index]['x']]
        z_coords = [df.loc[inner_toe_index]['z'], (df.loc[berm_inner_toe_side_index]['z'] + df.loc[berm_inner_crest_side_index]['z']) / 2, (df.loc[berm_inner_toe_side_index]['z'] + df.loc[berm_inner_crest_side_index]['z']) / 2, (df.loc[inner_crest_index]['z'] + df.loc[outer_crest_index]['z']) / 2, (df.loc[inner_crest_index]['z'] + df.loc[outer_crest_index]['z']) / 2, df.loc[outer_toe_index]['z']]

    # Plot the calculated profile coordinates, connected by a dotted line over the measured profile and the covariance of the profile over the cross-section
    fig, ax = plt.subplots(2, 1)
    ax[0].plot(profile_data['x'], profile_data['z'])
    ax[0].plot(df['x'], df['z_averaged'], 'r')
    ax[1].plot(df['x'], df['z_covariance'])
    ax[0].plot(x_coords, z_coords, color='k', marker='o', linestyle='--')
    ax[0].set_ylabel('m NAP')
    ax[1].set_ylabel('CoV profiel')
    ax[0].set_xlim(min(x_coords)-30, max(x_coords)+30)
    ax[1].set_xlim(min(x_coords)-30, max(x_coords)+30)
    fig.suptitle(titel)
    plt.savefig(path.joinpath(titel + '.png'))
    plt.close()
    return pd.DataFrame({'x': x_coords, 'z': z_coords})

def main():
    traject = '16-3'
    input_path = Path(r'c:\Users\wouterjanklerk\Documents\00_PhDGeneral\03_Cases\01_Rivierenland SAFE\WJKlerk\SAFE\data\InputFiles\Profiles').joinpath(traject)
    output_path = Path(r'c:\Users\wouterjanklerk\Documents\00_PhDGeneral\03_Cases\01_Rivierenland SAFE\WJKlerk\SAFE\data\InputFiles\Profiles').joinpath(traject,'profiles')
    input_file_name = 'InputProfiles.xlsx'
    output_filename = 'Dijkvakindeling_v5.2.xlsx'

    # Make output folder if not exist:
    if not output_path.is_dir():
        output_path.mkdir(parents=True, exist_ok=True)

    # Read the input fileconda
    input_data = pd.read_excel(input_path.joinpath(input_file_name))
    traject_data = pd.read_excel(input_path.parent.joinpath((output_filename)), header=1)
    traject_data = traject_data[(traject_data['Traject'] == traject) & (traject_data.iloc[:, 5])]

    # Extract profile information
    index = traject_data['dv_nummer'].reset_index(drop=True)

    for i in index:
        #Let op: extensies (bijv 34a) worden nu genegeerd!
        number = re.findall(r'\d+', str(i))
        extension = re.findall("[a-zA-Z]+", str(i))
        profile_number = 'Dwarsprofiel_' + str(number[0]).zfill(len(str(index.iloc[-1])))# if not extension else 'Dwarsprofiel_' + str(number[0]).zfill(len(str(index.iloc[-1]))) + str(extension[0])
        profile_data = input_data[input_data['dijkvaknummer'] == int(number[0])].reset_index(drop=True).rename(columns={'afstand buk [m, buitenkant +]': 'x', 'z_ahn [m NAP]': 'z', 'x': 'x_coord', 'y': 'y_coord'})
        profile = extract_profile_data(profile_data, window=2, titel=profile_number, path=output_path)

        # Save extracted data in .csv
        profile.to_csv(output_path.joinpath(profile_number + '.csv'))

        # Write profile_numbers to output file
        wb = load_workbook(input_path.parent.joinpath((output_filename)))
        ws = wb['Dijkvakindeling_keuze_info']

        for row in range(ws.max_row):
            if ws[row + 1][0].value == i:
                ws.cell(row=row + 1, column=11).value = profile_number
            else:
                pass

        wb.save(input_path.joinpath((output_filename)))

if __name__ == '__main__':
    main()
