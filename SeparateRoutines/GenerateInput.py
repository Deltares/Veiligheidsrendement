#This script makes the entire input structure from the general input files
import sys
sys.path.append('..')

import pandas as pd
from HydraRing_scripts import readDesignTable
from openpyxl import load_workbook
from pathlib import Path
from shutil import copyfile
from Mechanisms import OverflowSimple
import numpy as np
from scipy.optimize import fsolve
import shutil

def vka_measures(measures_basis, measures_custom_all, measures_location, vka_name, path, traject):
    j=len(measures_basis.index)
    measures_custom_vka1 =  measures_location.loc[vka_name].Value
    measures_basis.loc[j,'Name']=measures_custom_vka1
    measures_basis.loc[j,'Type']='Custom'
    measures_basis.loc[j,'File']= measures_custom_vka1 + '.csv'
    measures_basis.loc[j,'available']= 1   
    measures_basis.loc[j,'ID']= j + 1            
    #write custom to csv
    measure_custom_1 = measures_custom_all[measures_custom_all['variantnaam'] == measures_custom_vka1]
    naam= path.joinpath(traject, 'Output/Measures', measures_custom_vka1 + '.csv')
    measure_custom_1.to_csv(naam, index=False)
    measures_basis.loc[j, 'year'] = int(measure_custom_1.year-2025) #TODO, check 2025
    return(measures_basis)

def main():
    #TODO Somewhere in this function an extension should be made such that section specific information can also be inserted. Perhaps in separate files, named after the dike section.
    
    #Path of files. Should contain a subdirectory '\Input with designtables_HBN, designtables_TP, profiles, base_HBN.csv and measures.csv'
    # path = Path(r'c:\Users\wouterjanklerk\Documents\00_PhDGeneral\03_Cases\01_Rivierenland SAFE\WJKlerk\SAFE\data\InputFiles\Testcase_10sections_2021')
    path = Path(r'..\..\data\case_input\Testcase_10sections_2021')
    
    #Settings:
    traject = '16-4'                                                                            #Traject to consider
    file_name = 'Dijkvakindeling_v0.2.xlsx'                                                          #Name of main file
    backup_file_name = file_name + '.bak'                                                       #Name for backupping the main file before making changes
    fill_load_values = True                                                                     #If this is set to True, the script will fill missing values for crest height & temporal changes to loads from load_file.
                                                                                                # WARNING: this overwrites existing values!
    load_file = path.joinpath('Crest_WL_HBN_data.csv')                                          #File originating from the hydraulic load computations by HKV in 2018. Should be in same path as main file.
       
    #DO NOT USE: this changes the crest levels such that all overflow betas are in the given range.
    #TODO Put in separate routine.
    overflow_target_beta = False
    beta_t_overflow = [2.8,3.5]
    
    #Make a backup before adjusting the main file
    copyfile(path.joinpath(file_name), path.joinpath(backup_file_name))
    
    #Open and read data from Dijkvakindeling
    df = pd.read_excel(path.joinpath(file_name), sheet_name=None)
    
    #Adjust general sheet:
    DikeSections = df['Dijkvakindeling_keuze_info'].rename(columns=df['Dijkvakindeling_keuze_info'].iloc[0]).drop(df['Dijkvakindeling_keuze_info'].index[0])
    DikeSections = DikeSections[((DikeSections['Traject'] == traject) & (DikeSections['Wel of niet meerekenen'] == 1))].reset_index(drop=True)
    
    if 'FragilityCurve' in df['Info voor STBI'].columns.values:
        stbi_col = ['dwarsprofiel','FragilityCurve', 'SF_2025', 'SF_2075', 'd_cover','dSF/dberm', 'beta_2025', 'beta_2075', 'dbeta/dberm'] #new version excelfile
    else:
        stbi_col = ['dwarsprofiel', 'SF_2025', 'SF_2075', 'd_cover', 'dSF/dberm'] #old verion excelfile

    #Sheets for mechanisms:

    STBI_data = df['Info voor STBI'].loc[:,stbi_col]
    
    #Sheet for pinping:
    piping_col = ['dwarsprofiel','D','d70','d_cover','h_exit','r_exit','Lvoor','Lachter','k','gamma_sat','kwelscherm','dh_exit(t)']
    Piping_data = df['Info voor Piping'].loc[:, piping_col]
    
    #Sheet for housing:
    housing_col = ['Naam traject', 'Naam dijkvak', 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]    
    Housing = df['Info voor huizen'].loc[:, housing_col]
    
    #Sheet for measures:   
    if 'Maatregelen' in df.keys():
        measures_col = ['Naam traject','Naam dijkvak','Basismaatregelen','voorkeursalternatief_1','voorkeursalternatief_2','voorkeursalternatief_3']
        Measures = df['Maatregelen'].loc[:, measures_col]
    else: #old version without voorkeursalternatieven
        measures_per_location = pd.read_csv(path.joinpath(traject, 'Input/measures.csv'), delimiter=';')
    
    #If we want to fill missing load values based on the computations by HKV, it is done here:
    if fill_load_values:
        crest_heights = pd.read_csv(load_file, usecols=[0, 1], header=1)
        crest_heights['Dijkpaal'] = crest_heights['Dijkpaal'].str.replace('.', '')
    
        water_levels = pd.read_csv(load_file, usecols=[0,2,3,4,5,6,7], header=1)
        water_levels['Dijkpaal'] = water_levels['Dijkpaal'].str.replace('.', '')
    
        HBNs = pd.read_csv(load_file, usecols=[0,8,9,10,11,12,13], header=0,names = ['Dijkpaal', 2015, 2025, 2505, 2075, 2100, 2125])
        HBNs['Dijkpaal'] = HBNs['Dijkpaal'].str.replace('.', '')
        for i in DikeSections.index:
            #TODO improve this programming and dealing with headers etc (it's very sloppy and non-robust)
            df_CrestHeight = crest_heights[crest_heights['Dijkpaal'] == DikeSections['Dwarsprofiel piping'][i][:5]]
            CrestHeight = float(df_CrestHeight['Huidige Kruinhoogte [m+NAP]'])
    
            df_YearlyWLRise = water_levels[water_levels['Dijkpaal'] == DikeSections['Dwarsprofiel HoogteToetspeil'][i]]
            YearlyWLRise_factor = float((df_YearlyWLRise['2125'].values - df_YearlyWLRise['2015'].values) / (2125 - 2015))
    
            df_HBNRise_factor = HBNs[HBNs['Dijkpaal'] == DikeSections['Dwarsprofiel HoogteToetspeil'][i]]
            HBNRise_factor = float((df_HBNRise_factor[2125].values - df_HBNRise_factor[2015].values) / (2125 - 2015)) / YearlyWLRise_factor
    
            #Write YearlyWLRise and HBNRise factors to mastersheet
            wb = load_workbook(path.joinpath(file_name))
            ws = wb['Dijkvakindeling_keuze_info']
    
            for row in range(ws.max_row):
                if ws[row+1][8].value == DikeSections['Dwarsprofiel piping'][i] and ws[row+1][5].value == 1:
                    ws.cell(row=row+1, column=12).value = CrestHeight
                if ws[row+1][9].value == DikeSections['Dwarsprofiel HoogteToetspeil'][i] and ws[row+1][5].value == 1:
                    ws.cell(row=row+1, column=14).value = YearlyWLRise_factor
                    ws.cell(row=row+1, column=15).value = HBNRise_factor
                else:
                    pass
            wb.save(path.joinpath(file_name))
    
    #Check if two or multiple dike section are equally named
    if any(STBI_data['dwarsprofiel'].duplicated()) or any(Piping_data['dwarsprofiel'].duplicated()):
        raise Exception('Warning, two or multiple dike section are equally named!')
        sys.exit()
    
    #First we are going to write a general xlsx for every dike section considered:
    
    General = {}
    General['Name'] = ['Length', 'Start', 'End', 'Overflow', 'StabilityInner', 'Piping', 'LoadData', 'YearlyWLRise', 'HBNRise_factor']   
    
    #Make subfolders if not exist:
    if not path.joinpath(traject, 'Output').is_dir():
        path.joinpath(traject, 'Output/StabilityInner').mkdir(parents=True, exist_ok=True)
        path.joinpath(traject, 'Output/Piping').mkdir(parents=True, exist_ok=True)
        path.joinpath(traject, 'Output/Overflow').mkdir(parents=True, exist_ok=True)
        path.joinpath(traject, 'Output/Toetspeil').mkdir(parents=True, exist_ok=True)
    
    if overflow_target_beta:
        originalcrests= []
        newcrests= []
    
    for i in DikeSections.index:
    
        #Now we write subfiles with input for different submechanisms:
    
        #First we write files for StabilityInner:
        STBI_data_location = STBI_data[STBI_data['dwarsprofiel'] == DikeSections['Dwarsprofiel STBI/STBU'][i]].transpose().drop(['dwarsprofiel'], axis=0).reset_index()
        STBI_data_location.columns = ['Name', "Value"]
        STBI_data_location = STBI_data_location.set_index('Name')        
        if False if pd.isnull(STBI_data_location.loc['FragilityCurve', 'Value']) else True:
            General['Type'] = ['', '', '', 'Simple', 'FragilityCurve', 'SemiProb', '', '', '']
            src = path.joinpath(traject,'Input/FragilityCurve_STBI',STBI_data_location.loc['FragilityCurve', 'Value'])
            dst = path.joinpath(traject,'Output/FragilityCurve_STBI',STBI_data_location.loc['FragilityCurve', 'Value'])
            shutil.copy2(src, dst)
            #read fragilityCurve file
            # FC = pd.read_csv(path.joinpath(traject,'Input/FragilityCurve_STBI',STBI_data_location.loc['FragilityCurve', 'Value']), delimiter=';', header=0)
            # dA = pd.DataFrame([['h' ,FC['h'].values], ['beta', FC['beta'].values]], columns=['Name','Value']).set_index('Name')
            # STBI_data_location = STBI_data_location.append(dA)
        else: # uf fragility curve is not available
            General['Type'] = ['', '', '', 'Simple', 'Simple', 'SemiProb', '', '', '']
            if STBI_data_location['Value'][[1,2,5,6]].isnull().values.all():
                raise Exception('STBI data of cross-section {} (Dike section {}) contains NaN values'.format(DikeSections['Dwarsprofiel STBI/STBU'][i], DikeSections['dv_nummer'][i]))
                sys.exit()
        STBI_data_location.to_csv(path.joinpath(traject, 'Output/StabilityInner', DikeSections['Dwarsprofiel STBI/STBU'][i] + '_StabilityInner.csv'))
    
        #Then for piping:
        Piping_data_location = Piping_data[Piping_data['dwarsprofiel'] == DikeSections['Dwarsprofiel piping'][i]].transpose().drop(['dwarsprofiel'], axis=0).reset_index()
        if Piping_data_location.iloc[:, 1].isnull().values.any():
            raise Exception('Piping data of cross-section {} (Dike section {}) contains NaN values'.format(DikeSections['Dwarsprofiel STBI/STBU'][i], DikeSections['dv_nummer'][i]))
            sys.exit()
        Piping_data_location.columns = ['Name', "Value"]
        Piping_data_location = Piping_data_location.set_index('Name')
        Piping_data_location.to_csv(path.joinpath(traject, 'Output/Piping', DikeSections['Dwarsprofiel piping'][i] + '_Piping.csv'))
    
    
        #Then we read and write data for overflow (this is a bit more complicated):
        HBN_basis = pd.read_csv(path.joinpath(traject, 'Input/base_HBN.csv'), delimiter=';')
        OverflowData = readDesignTable(path.joinpath(traject, 'Input/designtables_HBN/DESIGNTABLE_' + DikeSections['Dwarsprofiel HoogteToetspeil'][i] + '.txt'))
    
        #Typically the design tables have 13 lines, but in some cases they dont. Then we should adjust the base file to have the right length:
        if len(OverflowData) != 13:
            HBN_basis = HBN_basis.iloc[0:len(OverflowData)]
        #Overwrite crestheight, betas, h_c and if possible dhc(t)
        if not overflow_target_beta:                            #This is the common variant. Quite straightforward
            HBN_basis['h_crest'].iloc[0] = DikeSections['Kruinhoogte'][i]
        else:                                                   #In this case the water levels are adjusted such that the reliability index is in a certain range. In principle, this is not used.
            beta_t = (beta_t_overflow[1] - beta_t_overflow[0]) * np.random.random_sample(1) + beta_t_overflow[0]
            crestlevel = fsolve(OverflowSimple,DikeSections['Kruinhoogte'][i],
                                args=(5, OverflowData['Value'].values, np.ones((len(OverflowData['Value'].values),1))*5,
                                      OverflowData['Beta\n'].values, 'assessment', None, None, True, beta_t))
            HBN_basis['h_crest'].iloc[0] = crestlevel[0]
            originalcrests.append(DikeSections['Kruinhoogte'][i])
            newcrests.append(crestlevel)
            print('for section ' + str(DikeSections.iloc[i]['dv_nummer']) + ' changed crest from ' + str(DikeSections['Kruinhoogte'][i]) + ' to ' + str(crestlevel))
        HBN_basis['h_c'] = OverflowData['Value'].values
        HBN_basis['beta'] = OverflowData['Beta\n'].values
    
        #If there is a value for crest level decrease available, put it in the sheet. This is not always the case, then a 0 is given. This is overridden with a default value in the main script.
        if len(DikeSections['Kruindaling'].value_counts()) > 0:
            HBN_basis['dhc(t)'].iloc[0] = DikeSections['Kruindaling'][i]
        else:
            HBN_basis['dhc(t)'].iloc[0] = 0.
        HBN_basis.to_csv(path.joinpath(traject, 'Output/Overflow', DikeSections['Dwarsprofiel HoogteToetspeil'][i] + '_Overflow.csv'), index=False)
    
        #Copy design tables for the water level:
        copyfile(path.joinpath(traject, 'Input/designtables_TP/DESIGNTABLE_' + DikeSections['Dwarsprofiel HoogteToetspeil'][i] + '.txt'), path.joinpath(traject, 'Output/Toetspeil/DESIGNTABLE_' + DikeSections['Dwarsprofiel HoogteToetspeil'][i] + '.txt'))
    
        #Fill general tab and write to Excel:
        General['Value'] = [DikeSections['Lengte dijkvak'][i],
                            DikeSections['Van'][i],
                            DikeSections['Tot'][i],
                            'Overflow/' + DikeSections['Dwarsprofiel HoogteToetspeil'][i] + '_Overflow.csv',
                            'StabilityInner/' + DikeSections['Dwarsprofiel STBI/STBU'][i] + '_StabilityInner.csv',
                            'Piping/' + DikeSections['Dwarsprofiel piping'][i] + '_Piping.csv',
                            'DESIGNTABLE_' + DikeSections['Dwarsprofiel HoogteToetspeil'][i] + '.txt',
                            DikeSections['Waterstandstijging'][i],
                            DikeSections['HBN factor'][i]]
        toExcel = pd.DataFrame.from_dict(General)[['Name', 'Value', 'Type']]
    
        if 'Maatregelen' in df.keys():
            #measures per section
            measures_location = Measures[Measures['Naam dijkvak'] == DikeSections['dv_nummer'][i]].transpose().drop(['Naam dijkvak'], axis=0).reset_index()
            measures_location.columns = ['Name', "Value"]
            measures_location = measures_location.set_index('Name')
            # read tables with measures
            measures_basis = pd.read_csv(path.joinpath(traject, 'Input/measures', measures_location.loc['Basismaatregelen'].Value), delimiter=';') 
            measures_custom_all = pd.read_csv(path.joinpath(traject, 'Input/measures/custom_measures.csv'))
            #add voorkeursalternatieven to table with measeares. 
            if (~pd.isna(measures_location.loc['voorkeursalternatief_1'].Value)== -1):
                measures_per_location = vka_measures(measures_basis, measures_custom_all, measures_location, 'voorkeursalternatief_1', path, traject)                   
    
            if (~pd.isna(measures_location.loc['voorkeursalternatief_2'].Value) == -1):
                measures_per_location = vka_measures(measures_basis, measures_custom_all, measures_location, 'voorkeursalternatief_2', path, traject) 
                               
            if (~pd.isna(measures_location.loc['voorkeursalternatief_3'].Value) == -1):
                measures_per_location = vka_measures(measures_basis, measures_custom_all, measures_location, 'voorkeursalternatief_3', path, traject)                  
           
    
        #Fill profile tab
        profile = pd.read_csv(path.joinpath(traject, 'Input/profiles', DikeSections['Dwarsprofiel Geometrie'][i] + '.csv'), index_col=0)
    
        #Fill houses tab
        houses_data_location = Housing[((Housing['Naam dijkvak'] == DikeSections['dv_nummer'][i]) & (Housing['Naam traject'] == traject))].transpose().drop(['Naam traject', 'Naam dijkvak'], axis=0).reset_index()
        houses_data_location.columns = ['distancefromtoe', 'number']
    
        #Write data
        try:
            writer = pd.ExcelWriter(path.joinpath(traject, 'Output/DV' + '{:02d}'.format(DikeSections['dv_nummer'][i])  + '.xlsx'))
        except:
            writer = pd.ExcelWriter(path.joinpath(traject, 'Output/DV' + DikeSections['dv_nummer'][i]  + '.xlsx'))
    
        toExcel.to_excel(writer, sheet_name='General', index=False)
        measures_per_location.to_excel(writer, sheet_name='Measures', index=False)
        profile.to_excel(writer, sheet_name='Geometry', index=False)
        houses_data_location.to_excel(writer, sheet_name='Housing', index=False)
        writer.save()


if __name__ == '__main__':
    main()
