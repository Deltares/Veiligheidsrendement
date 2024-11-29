from collections import defaultdict

import numpy as np

from scripts.design.combine_omega_length_effect import run_combination_single_database
from scripts.design.create_requirement_files import get_traject_requirements_per_sections, \
    create_requirement_files_lenient_and_strict, get_vr_eis_index_2075
from scripts.design.plot_design import plot_sensitivity

from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import pandas as pd
from openpyxl import Workbook

from scripts.postprocessing.database_access_functions import get_measures_for_run_id, get_measure_parameters, \
    get_measure_type
from scripts.postprocessing.database_analytics import get_measures_per_step_number, get_measures_per_section_for_step
from vrtool.common.enums import MechanismEnum


def read_sensitivity_results_and_plot(csv_result_path: Path, db_path):
    df_sensitivity = pd.read_csv(csv_result_path)
    _, vrm_optimization_steps, _, p_max = run_combination_single_database(db_path)
    plot_sensitivity(df_sensitivity, vrm_optimization_steps, p_max)
    return df_sensitivity


def compare_lenient_requirement_tables():
    path_dir = Path(r"C:\Users\hauth\OneDrive - Stichting Deltares\projects\VRTool\databases\38-1\automation")
    wb = Workbook()
    ws_base = wb.active
    ws_base.title = "Base Case"

    base_lenient_df = None
    highlight_fill = PatternFill(start_color="FFEE08", end_color="FFEE08", fill_type="solid")  # Yellow fill

    summary_data = []  # To collect counts of differences for the summary sheet
    params = {
        "name": "38-1",
        "eis": 1 / 10000,
        "l_traject": 28900,
        "revetment": False,
        "meet_eis": True,
    }

    for idx, db_path in enumerate(path_dir.glob("*.db")):
        print(db_path, "processing ...")
        params["path"] = db_path

        requirements_per_section = get_traject_requirements_per_sections(params)
        lenient_df, _ = create_requirement_files_lenient_and_strict(requirements_per_section, params)

        if idx == 0:
            # Assume first file is the base case
            base_lenient_df = lenient_df
            for r in dataframe_to_rows(base_lenient_df, index=False, header=True):
                ws_base.append(r)
            # Initialize summary data structure with 0 differences
            summary_data = [{"Row": i + 1, "OVERFLOW": 0, "PIPING": 0, "STABILITY_INNER": 0} for i in
                            range(len(base_lenient_df))]
        else:
            # Compare with the base case
            ws_sim = wb.create_sheet(title=f"Simulation {idx}")
            for r_idx, row in enumerate(dataframe_to_rows(lenient_df, index=False, header=True), start=1):
                ws_sim.append(row)
                if r_idx == 1:  # Skip header row for highlighting
                    continue
                for c_idx, col_name in enumerate(lenient_df.columns, start=1):
                    cell = ws_sim.cell(row=r_idx, column=c_idx)
                    base_value = base_lenient_df.iloc[r_idx - 2, c_idx - 1]  # Adjust for header row
                    current_value = lenient_df.iloc[r_idx - 2, c_idx - 1]
                    if col_name in ["OVERFLOW", "PIPING", "STABILITY_INNER"] and base_value != current_value:
                        cell.fill = highlight_fill  # Highlight cell with a different value
                        # Increment difference count for this row and column
                        summary_data[r_idx - 2][col_name] += 1

    # Add a summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    summary_df = pd.DataFrame(summary_data)
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(r)

    # Save workbook
    output_file = path_dir / "comparison_results_with_summary.xlsx"
    wb.save(output_file)
    print(f"Comparison saved to {output_file}")


def plot_sensitivity_box_plot_beta_all_runs(path_dir, mechanism: str):
    params = {
        "name": "38-1",
        "eis": 1 / 10000,
        "l_traject": 28900,
        "revetment": False,
        "meet_eis": True,
    }
    if mechanism == "piping":
        params["mechanism"] = MechanismEnum.PIPING
        filename = "beta_piping_per_section.csv"
    elif mechanism == "stability":
        params["mechanism"] = MechanismEnum.STABILITY_INNER
        filename = "beta_stability_per_section.csv"
    else:
        raise ValueError("Unknown mechanism")

    # betas are given for every section in a dicti : {1: 0.1, 2: 0.2, 3: 0.3, 4: 0.4}, I want to aggregate the betas per section for every database
    res = defaultdict(list)
    for idx, db_path in enumerate(path_dir.glob("*.db")):
        print(db_path, "processing ...")
        params["path"] = db_path
        requirements_per_section = get_traject_requirements_per_sections(params)

        piping_requirement_betas = requirements_per_section[params["mechanism"]]
        for section, beta in piping_requirement_betas.items():
            res[section].append(beta)

    # save dict to csv
    df = pd.DataFrame(res)
    df.to_csv(path_dir / filename)

    print(res)
    return df


def make_boxplot_sensitivity_per_sections(df):
    import matplotlib.pyplot as plt
    import seaborn as sns

    df = df.drop(columns="Unnamed: 0")  # remove first column
    data = df.to_numpy()  # convert df into nunpy array

    print(data)
    print(data.T[-1])

    fig = plt.figure(figsize=(10, 10))
    N = int(len(df))
    sns.boxplot(data, orient="h")
    plt.xlabel("Beta 2075")
    # rotate x labels
    plt.xticks(rotation=45)
    plt.ylabel("Sectie")
    plt.title(f"Boxplot of beta Piping per section- {N} runs")

    # save the plot
    plt.savefig("boxplot_beta_piping_per_section.png")
    plt.show()


def create_comparison_section_table(path_dir: Path, run_id: int, path_base_db: Path = None):
    """
    The base case database MUST be:
        - in the path_dir
        - be the first database

    Args:
        path_dir:
        run_id:

    Returns:

    """
    # iterate over all databases in the directory
    all_data = []

    # Get data
    for idx, db_path in enumerate(path_dir.glob("*.db")):

        print(db_path, "processing ...")
        if idx == 0:
            run_idd = 1 # force run_id for base case
            assert db_path.stem == "38-1_basis_0", "Verify that first iterated database is the base case"
        else:
            run_idd = run_id

        params = {
            "name": "38-1",
            "eis": 1 / 10000,
            "l_traject": 28900,
            "revetment": False,
            "meet_eis": True,
            "path": db_path
        }
        step_requirement: int = get_vr_eis_index_2075(params)
        lists_of_measures = get_measures_for_run_id(db_path, run_id=run_idd)

        measures_per_step = get_measures_per_step_number(lists_of_measures)

        measures_per_section = get_measures_per_section_for_step(
            measures_per_step, step_requirement + 1
        )

        def _set_section_parameters(db_location: Path):
            run_data = {}
            for section in measures_per_section:
                section_data = {"name": [], "params": {"dberm": None, "dcrest": None, "L_screen": None}}
                for measure in measures_per_section[section][0]:
                    measure_params = get_measure_parameters(measure, db_location)
                    section_data["params"]["dberm"] = measure_params.get("dberm")
                    section_data["params"]["dcrest"] = measure_params.get("dcrest")
                    section_data["params"]["L_screen"] = measure_params.get("l_stab_screen")
                    measure_name = get_measure_type(measure, db_location)
                    section_data["name"].append(measure_name.get("name"))

                section_data["name"] = "".join(section_data["name"])  # convert list to string
                run_data[section] = section_data
            return run_data

        all_data.append(_set_section_parameters(db_path))

    # data base case
    base_case = all_data[0]
    count_diff_meas_type = np.zeros(len(base_case))
    count_diff_dberm = np.zeros(len(base_case))
    count_diff_dcrest = np.zeros(len(base_case))
    count_diff_L_screen = np.zeros(len(base_case))
    vzg_present = np.zeros(len(base_case))
    stab_screen_present = np.zeros(len(base_case))
    damwand_present = np.zeros(len(base_case))

    dim_array = (len(base_case), len(all_data))
    array_dberm = np.zeros(dim_array)
    array_dcrest = np.zeros(dim_array)
    array_L_screen = np.zeros(dim_array)


    # Process data
    for i, run in enumerate(all_data):
        for j, section in enumerate(run):
            if run[section]["name"] != base_case[section]["name"]:
                count_diff_meas_type[j] += 1
            if run[section]["params"]["dberm"] != base_case[section]["params"]["dberm"]:
                count_diff_dberm[j] += 1
            if run[section]["params"]["dcrest"] != base_case[section]["params"]["dcrest"]:
                count_diff_dcrest[j] += 1
            if run[section]["params"]["L_screen"] != base_case[section]["params"]["L_screen"]:
                count_diff_L_screen[j] += 1

            # NOT ROBUST TO RELY ON THE USER NAME PROVIDED FOR THE MEASURE!
            if "Verticaal Zanddicht" in run[section]["name"]:
                vzg_present[j] += 1
            if "stabiliteitsscherm" in run[section]["name"]:
                stab_screen_present[j] += 1
            if "Zelfkerende constructie" in run[section]["name"]:
                damwand_present[j] += 1

            # get the parameters
            array_dberm[j, i] = run[section]["params"]["dberm"]
            array_dcrest[j, i] = run[section]["params"]["dcrest"]
            array_L_screen[j, i] = run[section]["params"]["L_screen"]

    mean_dberm = np.mean(array_dberm, axis=1)
    mean_dcrest = np.mean(array_dcrest, axis=1)
    mean_l_stab_screen = np.mean(array_L_screen, axis=1)

    count_unique_dcrest = np.array([len(np.unique(array_dcrest[i])) for i in range(len(base_case))])
    count_unique_dberm = np.array([len(np.unique(array_dberm[i])) for i in range(len(base_case))])
    count_unique_Lscreen = np.array([len(np.unique(array_L_screen[i])) for i in range(len(base_case))])

    # get info on the base case:
    print("Base case:")
    base_measure_type = np.array([base_case[section]["name"] for section in base_case])
    base_dberm = np.array([base_case[section]["params"]["dberm"] for section in base_case])
    base_dcrest = np.array([base_case[section]["params"]["dcrest"] for section in base_case])
    base_L_screen = np.array([base_case[section]["params"]["L_screen"] for section in base_case])


    # put data in a csv file
    df = pd.DataFrame({
        "Section name": list(base_case.keys()),
        "Base_measure_type": base_measure_type,
        "Base_dberm": base_dberm,
        "Base_dcrest": base_dcrest,
        "Base_L_screen": base_L_screen,
        "Diff_meas_type": count_diff_meas_type,
        "Diff_dberm": count_diff_dberm,
        "Diff_dcrest": count_diff_dcrest,
        "Diff_L_screen": count_diff_L_screen,
        "VZG_present": vzg_present,
        "Stab_screen_present": stab_screen_present,
        "Damwand_present": damwand_present,
        "mean_dberm": mean_dberm,
        "mean_dcrest": mean_dcrest,
        "mean_L_stab_screen": mean_l_stab_screen,
        "count_unique_dberm": count_unique_dberm,
        "count_unique_dcrest": count_unique_dcrest,
        "count_unique_Lscreen": count_unique_Lscreen
    })

    df.to_csv(path_dir / "comparison_section_table.csv")
    #save to excel
    df.to_excel(path_dir / "comparison_section_table.xlsx", index=False)
    return all_data